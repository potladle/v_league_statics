from volleyball.stat.StatMaker import *
from typing import List
import openpyxl

"""js = '''{"name":"Cristian", "address":{"street":"Sesame","number":122}}'''
    j = json.loads(js)
    print(j)
    u = User(**j)
    print(u)

two_messages: List[Messages] = [Messages([]), Messages([])]
    with open(path, encoding="UTF-8") as file:
        json = json.load(file)

        for i, messages in enumerate(json):
            for one_message_json in messages:
                if one_message_json["action"] == "팀득점" or one_message_json["action"] == "팀실패":
                    two_messages[i].append(TeamMessage(one_message_json["index"], one_message_json["success_failure"]))
                else:
                    two_messages[i].append(action_to_class[one_message_json["action"]](**remove_key(remove_key(one_message_json, "action"), "error_type")))
                two_messages[i][last_index(two_messages[i])].error_type = one_message_json["error_type"]

    for messages in two_messages:
        for message in messages:
            print(message.__dict__, message.__class__)

        print("--------------------------------------")
"""


stat_file = StatFile("C:\\Users\\jsmoo\\OneDrive\\바탕 화면\\배구", ["여자부"], [2021], list(range(1, 300)))
stat_file.fill_games()
stat_file.fill_sets()
stat_file.fill_messages()
stat_file.fill_rallies()
stat_file.fill_chances()


class AttackTypePlayer(Player):

    def __init__(self, name: str, back_num: int, season: int, team: str, attack_type_in_class: str):
        super().__init__(name, back_num, season, team)
        self.attack_type_in_class = attack_type_in_class
        self.team = team


attack_type_eff : EffCalc = EffCalc({})


def processor_in_set(messages: List[Messages], chances: Chances, rallies: Rallies, game: Game, set_in_game: Set):
    attack_messages = messages[0].select_messages_with_action("공격종합")
    for message in attack_messages:
        attack_type_eff.append_effect_calc(Messages([message]), game, "공격종합", AttackTypePlayer, take_key_by_value(action_to_class, type(message)))

    attack_messages = messages[1].select_messages_with_action("공격종합")
    for message in attack_messages:
        attack_type_eff.append_effect_calc(Messages([message]), game, "공격종합", AttackTypePlayer, take_key_by_value(action_to_class, type(message)))


stat_file.process_in_set(processor_in_set)


wb = openpyxl.Workbook()

for i, one_team in enumerate(women_teams):
    wb.create_sheet(one_team, index=i)
    sheet = wb[one_team]
    sheet.cell(row=1, column=1).value = "선수"
    attack_type_effects: List[EffCalc] = []

    for v, attack_type in enumerate(attack_types):
        sheet.cell(row=1, column=v+2).value = attack_type
        sheet.cell(row=1, column=v+9).value = attack_type + " 점유율"
        attack_type_effects.append(EffCalc(filter(lambda item: item[0].attack_type_in_class == attack_type and item[0].team == one_team, attack_type_eff.items())))

    sheet.cell(row=1, column=8).value = "총 공격 시도"

    for v, player in enumerate(attack_type_effects[0].keys()):
        sheet.cell(row=v+2, column=1).value = player.name
        trials = map(lambda attack_type_effect: attack_type_effect.find_stat_with_name(player.name)[0], attack_type_effects)

        whole = 0

        for index, trial in enumerate(trials):
            sheet.cell(row=v+2, column=index+2).value = trial
            whole += trial

        trials = map(lambda attack_type_effect: attack_type_effect.find_stat_with_name(player.name)[0], attack_type_effects)

        sheet.cell(row=v+2, column=8).value = whole

        for index, trial in enumerate(trials):
            print(v+2, index+8)
            sheet.cell(row=v+2, column=index+9).value = f"{make_percentage(whole, trial)}%"

wb.save("공격 종류.xlsx")

wb.close()
