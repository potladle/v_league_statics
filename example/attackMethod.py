from volleyball.stat.StatMaker import *
from typing import List
import openpyxl

stat_file = StatFile("C:\\Users\\jsmoo\\OneDrive\\바탕 화면\\배구", ["여자부"], [2021], list(range(1, 300)))
stat_file.fill_games()
stat_file.fill_sets()
stat_file.fill_messages()
stat_file.fill_rallies()
stat_file.fill_chances()


class AttackMethodPlayer(Player):
    COURT_IN = 0
    BLOCK_OUT = 1
    DIG_FAULT = 2

    def __init__(self
                 , name: str
                 , back_num: int
                 , season: int
                 , team: str
                 , court_in_or_block_out: int
                 , receive_result):
        super(AttackMethodPlayer, self).__init__(name, back_num, season, team)
        self.attack_method = court_in_or_block_out
        self.receive_result = receive_result


attack_method_eff : EffCalc = EffCalc({})


def processor_in_set(messages: List[Messages], chances: Chances, rallies: Rallies, game: Game, set_in_game: Set):
    attack_messages = messages[0].select_messages_with_action("공격종합")
    for attack_message in attack_messages:
        receive_result = 0
        chance = chances.select_chance_with_message_index(attack_message.index)
        if chance.index == len(chances) - 1 or chance.side == 1:
            continue

        first_message_chance = messages[0].slice_messages_with_chance(chance)[0]
        next_opposite_first_message_chance = messages[1].slice_messages_with_chance(chances[chance.index + 1])[0]

        if type(next_opposite_first_message_chance) == TeamMessage and next_opposite_first_message_chance.success_failure == Message.FAILURE:
            message_attack_method = AttackMethodPlayer.COURT_IN
        elif type(next_opposite_first_message_chance) == BlockMessage and next_opposite_first_message_chance.success_failure == Message.FAILURE:
            message_attack_method = AttackMethodPlayer.BLOCK_OUT
        elif type(next_opposite_first_message_chance) == DigMessage and next_opposite_first_message_chance.success_failure == Message.FAILURE:
            message_attack_method = AttackMethodPlayer.DIG_FAULT
        else:
            continue

        if first_message_chance.is_right(ReceiveMessage, Message.SUCCESS):
            continue

        attack_method_eff.append_effect_calc(Messages([attack_message]), game, "공격종합", AttackMethodPlayer, message_attack_method, receive_result)

    attack_messages = messages[1].select_messages_with_action("공격종합")
    for attack_message in attack_messages:
        receive_result = 0
        chance = chances.select_chance_with_message_index(attack_message.index)
        if chance.index == len(chances) - 1 or chance.side == 0:
            continue

        first_message_chance = messages[1].slice_messages_with_chance(chance)[0]
        next_opposite_first_message_chance = messages[0].slice_messages_with_chance(chances[chance.index + 1])[0]

        if type(next_opposite_first_message_chance) == TeamMessage and next_opposite_first_message_chance.success_failure == Message.FAILURE:
            message_attack_method = AttackMethodPlayer.COURT_IN
        elif type(next_opposite_first_message_chance) == BlockMessage and next_opposite_first_message_chance.success_failure == Message.FAILURE:
            message_attack_method = AttackMethodPlayer.BLOCK_OUT
        elif type(next_opposite_first_message_chance) == DigMessage and next_opposite_first_message_chance.success_failure == Message.FAILURE:
            message_attack_method = AttackMethodPlayer.DIG_FAULT
        else:
            continue

        if first_message_chance.is_right(ReceiveMessage, Message.SUCCESS):
            continue

        attack_method_eff.append_effect_calc(Messages([attack_message]), game, "공격종합", AttackMethodPlayer, message_attack_method, receive_result)


stat_file.process_in_set(processor_in_set)

attack_method_eff.print_effect_with_name(top_n=1000)

wb = openpyxl.Workbook()

for i, one_team in enumerate(women_teams):
    wb.create_sheet(one_team, index=i)
    sheet = wb[one_team]

    court_in_eff = EffCalc(filter(lambda item: item[0].attack_method == AttackMethodPlayer.COURT_IN and item[0].team == one_team, attack_method_eff.items()))
    block_out_eff = EffCalc(filter(lambda item: item[0].attack_method == AttackMethodPlayer.BLOCK_OUT and item[0].team == one_team, attack_method_eff.items()))
    dig_fault_eff = EffCalc(filter(lambda item: item[0].attack_method == AttackMethodPlayer.DIG_FAULT and item[0].team == one_team, attack_method_eff.items()))

    court_in_eff.print_effect_with_name(prefix=f"{one_team} 코트 인")
    block_out_eff.print_effect_with_name(prefix=f"{one_team} 블로킹 터치 아웃")
    dig_fault_eff.print_effect_with_name(prefix=f"{one_team} 디그 실패")

    variables = ["선수", "코트 인", "디그 실패", "블로킹 터치 아웃", "전체 공격 성공", "코트 안 공격"]
    for v, variable in enumerate(variables):
        sheet.cell(row=1, column=v + 1).value = variable

    for v, player in enumerate(court_in_eff.keys()):
        court_in_attack = court_in_eff.find_stat_with_name(player.name)[0]
        dig_fault_attack = dig_fault_eff.find_stat_with_name(player.name)[0]
        block_out_attack = block_out_eff.find_stat_with_name(player.name)[0]

        sheet.cell(row=v + 2, column=1).value = player.name
        sheet.cell(row=v + 2, column=2).value = court_in_attack
        sheet.cell(row=v + 2, column=3).value = dig_fault_attack
        sheet.cell(row=v + 2, column=4).value = block_out_attack
        sheet.cell(row=v + 2, column=5).value = court_in_attack + dig_fault_attack + block_out_attack
        sheet.cell(row=v + 2, column=6).value = f"{make_percentage(court_in_attack + dig_fault_attack + block_out_attack, court_in_attack + dig_fault_attack)}%"

wb.save("공격 성공.xlsx")

wb.close()
