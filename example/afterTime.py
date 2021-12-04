from statistics import mean

from volleyball.stat.StatMaker import *
from typing import List
from openpyxl import Workbook

stat_file = StatFile("C:\\Users\\jsmoo\\OneDrive\\바탕 화면\\배구", ["여자부"], [2021], list(range(1, 300)))
stat_file.fill_games()
stat_file.fill_sets()
stat_file.fill_messages()
stat_file.fill_rallies()
stat_file.fill_chances()

time_after_win: Dict[str, List[int]] = {}

for team in women_teams:
    time_after_win[team] = []

attack_after_win: Dict[str, Dict[str, List[int]]] = {}

end_actions = ["오픈", "퀵오픈", '백어택', '속공', '이동', '시간차', '블로킹', '서브']

for end_action in end_actions:
    attack_after_win[end_action] = {}
    for team in women_teams:
        attack_after_win[end_action][team] = []


def processor_in_set(messages: List[Messages], chances: Chances, rallies: Rallies, game: Game, set_in_game: Set):
    time_messages: Messages = Messages([])
    for message in messages[0]:
        if type(message) == TimeMessage:
            time_messages.append(message)

    for time_message in time_messages:
        after_rally = rallies.select_rally_after_message_index(time_message.index)[0]
        win_persistence = 0
        try:
            if after_rally.winner == Rally.RIGHT:
                win_persistence = -1
                while True:
                    if rallies[rallies.index(after_rally) + abs(win_persistence)].winner == Rally.RIGHT:
                        win_persistence -= 1
                    else:
                        time_after_win[game.teams[Rally.LEFT]].append(win_persistence)
                        break
            else:
                win_persistence = 1
                while True:
                    if rallies[rallies.index(after_rally) + abs(win_persistence)].winner == Rally.LEFT:
                        win_persistence += 1
                    else:
                        time_after_win[game.teams[Rally.LEFT]].append(win_persistence)
                        break
        except IndexError:
            time_after_win[game.teams[Rally.LEFT]].append(win_persistence)

    time_messages: Messages = Messages([])
    for message in messages[Rally.RIGHT]:
        if type(message) == TimeMessage:
            time_messages.append(message)

    for time_message in time_messages:
        after_rally = rallies.select_rally_after_message_index(time_message.index)[0]
        win_persistence = 0
        try:
            if after_rally.winner == Rally.LEFT:
                win_persistence = -1
                while True:
                    if rallies[rallies.index(after_rally) + abs(win_persistence)].winner == Rally.LEFT:
                        win_persistence -= 1
                    else:
                        time_after_win[game.teams[Rally.RIGHT]].append(win_persistence)
                        break
            else:
                print(1)
                win_persistence = 1
                while True:
                    if rallies[rallies.index(after_rally) + abs(win_persistence)].winner == Rally.RIGHT:
                        win_persistence += 1
                    else:
                        time_after_win[game.teams[Rally.RIGHT]].append(win_persistence)
                        break
        except IndexError:
            time_after_win[game.teams[Rally.RIGHT]].append(win_persistence)

    for message in messages[0]:
        if type(message) in list(map(lambda one_action: action_to_class[one_action], end_actions)) and message.success_failure == Message.SUCCESS:
            rally = rallies.select_rally_inner_message_index(message.index)
            rally_index = rallies.index(rally)
            win_persistence = 0
            try:
                while True:
                    if rallies[rally_index + 1 + abs(win_persistence)].winner == Rally.LEFT:
                        if win_persistence >= 0:
                            win_persistence += 1
                        else:
                            attack_after_win[take_key_by_value(action_to_class, type(message))][game.teams[0]].append(win_persistence)
                            break
                    else:
                        if win_persistence <= 0:
                            win_persistence -= 1
                        else:
                            attack_after_win[take_key_by_value(action_to_class, type(message))][game.teams[0]].append(win_persistence)
                            break
            except IndexError:
                attack_after_win[take_key_by_value(action_to_class, type(message))][game.teams[0]].append(win_persistence)

    for message in messages[1]:
        if type(message) in list(map(lambda one_action: action_to_class[one_action], end_actions)) and message.success_failure == Message.SUCCESS:
            rally = rallies.select_rally_inner_message_index(message.index)
            rally_index = rallies.index(rally)
            win_persistence = 0
            try:
                while True:
                    if rallies[rally_index + 1 + abs(win_persistence)].winner == Rally.RIGHT:
                        if win_persistence >= 0:
                            win_persistence += 1
                        else:
                            attack_after_win[take_key_by_value(action_to_class, type(message))][game.teams[1]].append(win_persistence)
                            break
                    else:
                        if win_persistence <= 0:
                            win_persistence -= 1
                        else:
                            attack_after_win[take_key_by_value(action_to_class, type(message))][game.teams[1]].append(win_persistence)
                            break
            except IndexError:
                attack_after_win[take_key_by_value(action_to_class, type(message))][game.teams[1]].append(win_persistence)


stat_file.process_in_set(processor_in_set)

wb = Workbook()
ws = wb.active

ws.cell(row=1, column=2).value = "타임 이후"

for v, end_action in enumerate(end_actions):
    ws.cell(row=1, column=3+v).value = end_action

for team, time_after_win_rally in remove_key(time_after_win, "페퍼저축은행").items():
    ws.cell(row=list(time_after_win.keys()).index(team) + 2, column=1).value = team

    ws.cell(row=list(time_after_win.keys()).index(team) + 2, column=2).value = mean(time_after_win_rally)

    for action, after_win in attack_after_win.items():
        ws.cell(row=list(after_win.keys()).index(team) + 2, column=list(attack_after_win.keys()).index(action) + 3).value = mean(after_win[team])


wb.save("타임 이후 연속 득점.xlsx")
